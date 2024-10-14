from openpyxl import Workbook, load_workbook
#------------------------------------------
wb = Workbook();      # guardar un instancia
wb.save("balance.xlsx")  # guardar libro en excel: CREAR
#------------------------------------------
# importar libro: cargar un excel
wb = load_workbook(r"C:\xampp\htdocs\PHP_PYTHON_CHATGPT\pruebas\manejar_openpyxl\balance.xlsx", read_only=False)
ws = wb.active     # cargar datos: simpre te devuelve la primera hoja
#------------------------------------------
ws1 = wb.create_sheet('myhoja1')      # crear hojas en excel
ws2 = wb.create_sheet('myhoja2',0)    # creo y sañalas que posicion tiene
ws2.title = 'nueva_hoja2'     # cambiar nombre a la hoja
wb.save(r"C:\xampp\htdocs\PHP_PYTHON_CHATGPT\pruebas\manejar_openpyxl\balance.xlsx") #guardar cambios
#------------------------------------------
# Copear contenido de una hoja a otra hoja// copea la que estamos usando o la activa
wb.sheetname  #obetenemos el nombre de las hojas
ws3 = wb['myhoja1']
fuente = wb.active
nueva_hoja = wb.copy_worksheet(fuente)
wb.remove_sheet(ws3)
wb.save(r"C:\xampp\htdocs\PHP_PYTHON_CHATGPT\pruebas\manejar_openpyxl\balance.xlsx")
#------------------------------------------
# Acceder y llenar en la celda especifica excel
wb.sheetnames
ws = wb['Sheet']  #elejimos la hoa Sheet
c = ws['A4'].value = 155 #almacenar 155
d = ws.cell(row=10, column = 8, value = " mario")  #ubicacion mas precisa
wb.save(r"C:\xampp\htdocs\PHP_PYTHON_CHATGPT\pruebas\manejar_openpyxl\balance.xlsx")
print(c)
#------------------------------------------
# llenado de manera randon
import random
wb.sheetnames
ws = wb['Sheet']  #elejimos la hoja Sheet
n = 10
dato = ['mario','juan', 'pedro']
for x in range(n):
    nombre = dato[random.randint(0,len(dato)-1)]
    ws.cell(column=1, row=x+1, value=nombre)
    ws.cell(column=2, row=x+1, value=random.randint(23,89))
wb.save(r"C:\xampp\htdocs\PHP_PYTHON_CHATGPT\pruebas\manejar_openpyxl\balance.xlsx")